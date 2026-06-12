"""Excel (.xlsx) tools for JOTARO's --workspace mode.

Same shape as ai/fs_tools.py: reads/lists are free, mutations (write cell / append row
/ create) are in EXCEL_WRITE_TOOLS so the harness asks the user to confirm first. All
paths go through the SAME workspace path-jail (`fs_tools._safe`). Tool names are
prefixed `excel_` so the combined workspace dispatcher can route them.

Backed by openpyxl. `make_excel_dispatcher(base)` returns dispatch(name, args).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Callable

import openpyxl

from ai.fs_tools import PathEscape, _safe

log = logging.getLogger(__name__)

MAX_ROWS = 100  # cap rows returned by excel_read (keep prompts small)

EXCEL_WRITE_TOOLS: frozenset[str] = frozenset(
    {"excel_write_cell", "excel_append_row", "excel_create"}
)

EXCEL_TOOL_SPECS: list[dict[str, Any]] = [
    {
        "type": "function",
        "function": {
            "name": "excel_list_sheets",
            "description": "List the sheet (tab) names in an .xlsx workbook.",
            "parameters": {
                "type": "object",
                "properties": {"path": {"type": "string", "description": "workbook path in the workspace"}},
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_read",
            "description": "Read rows from a sheet of an .xlsx workbook (returns a grid of values).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "description": "sheet name (default: the first/active sheet)"},
                    "max_rows": {"type": "integer", "description": f"rows to return (default {MAX_ROWS})"},
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_write_cell",
            "description": "Set ONE cell's value in a sheet, e.g. cell 'B2' (needs confirmation).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string"},
                    "cell": {"type": "string", "description": "A1-style reference, e.g. 'C5'"},
                    "value": {"description": "the value to write (text or number)"},
                },
                "required": ["path", "cell", "value"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_append_row",
            "description": "Append a row of values to the end of a sheet (needs confirmation).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string"},
                    "values": {"type": "array", "items": {}, "description": "the cells of the new row, left to right"},
                },
                "required": ["path", "values"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_create",
            "description": "Create a NEW empty .xlsx workbook (errors if the file already exists; needs confirmation).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "description": "name of the first sheet (default 'Sheet1')"},
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_find_rows",
            "description": (
                "Find rows where a COLUMN equals a value (the sheet's first row is treated "
                "as headers). Use this to answer 'who/which rows have X' — e.g. column='dept', "
                "equals='HR'. Do NOT pass a column value as a sheet name."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "description": "sheet name (default: first sheet)"},
                    "column": {"type": "string", "description": "header name to match on"},
                    "equals": {"description": "value the column must equal (case-insensitive)"},
                },
                "required": ["path", "column", "equals"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_aggregate",
            "description": "Compute sum/avg/min/max/count over a numeric COLUMN (by header name).",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string"},
                    "column": {"type": "string", "description": "header name of the column"},
                    "op": {"type": "string", "enum": ["sum", "avg", "min", "max", "count"]},
                },
                "required": ["path", "column", "op"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_read_range",
            "description": "Read a specific A1-style cell range, e.g. 'A1:C10'.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string"},
                    "range": {"type": "string", "description": "A1-style range, e.g. 'A1:C10'"},
                },
                "required": ["path", "range"],
            },
        },
    },
]


def _pick_sheet(wb: Any, name: str | None) -> Any:
    if not name:
        return wb.active
    if name not in wb.sheetnames:
        raise KeyError(name)
    return wb[name]


def _load_table(p: Any, sheet_name: str | None) -> tuple[str, list[str], list[list]]:
    """Return (sheet_title, headers, data_rows) — the first row is treated as headers."""
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb, sheet_name)
        title = ws.title
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    headers = [("" if c is None else str(c)) for c in rows[0]] if rows else []
    return title, headers, rows[1:]


def _col_index(headers: list[str], column: str) -> int:
    target = str(column).strip().lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == target:
            return i
    return -1


def make_excel_dispatcher(base_dir: str | Path) -> Callable[[str, dict[str, Any]], Any]:
    base = Path(base_dir).resolve()
    base.mkdir(parents=True, exist_ok=True)

    def dispatch(name: str, args: dict[str, Any]) -> Any:
        args = args or {}
        try:
            if name == "excel_list_sheets":
                p = _safe(base, str(args.get("path", "")))
                wb = openpyxl.load_workbook(p, read_only=True)
                try:
                    return {"path": str(p.relative_to(base)), "sheets": wb.sheetnames}
                finally:
                    wb.close()

            if name == "excel_read":
                p = _safe(base, str(args.get("path", "")))
                try:
                    limit = max(1, min(int(args.get("max_rows", MAX_ROWS)), 1000))
                except (TypeError, ValueError):
                    limit = MAX_ROWS
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                try:
                    ws = _pick_sheet(wb, args.get("sheet"))
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= limit:
                            break
                        rows.append([("" if c is None else c) for c in row])
                    return {"path": str(p.relative_to(base)), "sheet": ws.title,
                            "headers": rows[0] if rows else [],
                            "rows": rows, "truncated": ws.max_row > limit}
                finally:
                    wb.close()

            if name == "excel_write_cell":
                p = _safe(base, str(args.get("path", "")))
                wb = openpyxl.load_workbook(p)
                ws = _pick_sheet(wb, args.get("sheet"))
                ws[str(args["cell"])] = args.get("value")
                wb.save(p)
                log.warning("xlsx WRITE excel_write_cell %s %s=%r", p, args.get("cell"), args.get("value"))
                return {"status": "written", "path": str(p.relative_to(base)),
                        "sheet": ws.title, "cell": str(args["cell"])}

            if name == "excel_append_row":
                p = _safe(base, str(args.get("path", "")))
                values = args.get("values") or []
                if not isinstance(values, list):
                    return {"error": "values must be a list"}
                wb = openpyxl.load_workbook(p)
                ws = _pick_sheet(wb, args.get("sheet"))
                ws.append(values)
                wb.save(p)
                log.warning("xlsx WRITE excel_append_row %s (%d cells)", p, len(values))
                return {"status": "appended", "path": str(p.relative_to(base)),
                        "sheet": ws.title, "cells": len(values)}

            if name == "excel_create":
                p = _safe(base, str(args.get("path", "")))
                if p.exists():
                    return {"error": "file already exists — use excel_write_cell/append instead"}
                p.parent.mkdir(parents=True, exist_ok=True)
                wb = openpyxl.Workbook()
                wb.active.title = str(args.get("sheet") or "Sheet1")
                wb.save(p)
                log.warning("xlsx WRITE excel_create %s", p)
                return {"status": "created", "path": str(p.relative_to(base))}

            if name == "excel_find_rows":
                p = _safe(base, str(args.get("path", "")))
                title, headers, data = _load_table(p, args.get("sheet"))
                idx = _col_index(headers, str(args.get("column", "")))
                if idx < 0:
                    return {"error": f"column not found; available headers: {headers}"}
                want = str(args.get("equals", "")).strip().lower()
                matches = []
                for row in data:
                    cell = row[idx] if idx < len(row) else None
                    if str("" if cell is None else cell).strip().lower() == want:
                        matches.append({h: ("" if v is None else v)
                                        for h, v in zip(headers, row)})
                        if len(matches) >= 100:
                            break
                return {"sheet": title, "column": args.get("column"),
                        "equals": args.get("equals"), "matches": matches, "count": len(matches)}

            if name == "excel_aggregate":
                p = _safe(base, str(args.get("path", "")))
                title, headers, data = _load_table(p, args.get("sheet"))
                idx = _col_index(headers, str(args.get("column", "")))
                if idx < 0:
                    return {"error": f"column not found; available headers: {headers}"}
                op = str(args.get("op", "count")).lower()
                col = [row[idx] for row in data if idx < len(row) and row[idx] is not None]
                nums = [float(v) for v in col if isinstance(v, (int, float))]
                if op == "count":
                    val: Any = len(col)
                elif not nums:
                    return {"error": "no numeric values in that column"}
                elif op == "sum":
                    val = sum(nums)
                elif op == "avg":
                    val = sum(nums) / len(nums)
                elif op == "min":
                    val = min(nums)
                elif op == "max":
                    val = max(nums)
                else:
                    return {"error": f"unknown op {op!r}"}
                return {"sheet": title, "column": args.get("column"), "op": op,
                        "value": val, "n": len(col)}

            if name == "excel_read_range":
                p = _safe(base, str(args.get("path", "")))
                rng = str(args.get("range", ""))
                wb = openpyxl.load_workbook(p, data_only=True)
                try:
                    ws = _pick_sheet(wb, args.get("sheet"))
                    cells = ws[rng]
                    rows = []
                    if not isinstance(cells, tuple):       # single cell
                        rows = [["" if cells.value is None else cells.value]]
                    else:
                        for r in cells:
                            if isinstance(r, tuple):
                                rows.append([("" if c.value is None else c.value) for c in r])
                            else:
                                rows.append(["" if r.value is None else r.value])
                    return {"sheet": ws.title, "range": rng, "rows": rows}
                finally:
                    wb.close()

            return {"error": f"unknown tool {name!r}"}
        except PathEscape as exc:
            return {"error": str(exc)}
        except KeyError as exc:
            return {"error": f"sheet not found: {exc}"}
        except FileNotFoundError:
            return {"error": "workbook not found"}
        except Exception as exc:  # openpyxl raises various — surface, don't crash the loop
            return {"error": f"{type(exc).__name__}: {exc}"}

    return dispatch
